import openpyxl as xl
wb = xl.load_workbook('./transaction.xlsx')
sheet = wb['Sheet1']
cell = sheet.cell(1,1)
print(cell.value)
print(sheet.max_row)